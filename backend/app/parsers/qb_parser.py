"""
QuickBooks .xlsx export parser.

Pure function — no DB access, no FastAPI imports.
Input: raw file bytes + filename string
Output: structured dict ready for auto-mapping and storage

Supports both Profit & Loss and Balance Sheet exports.
Multi-month exports (e.g. Jan–Dec columns) are handled naturally.
"""
import io
from decimal import Decimal, ROUND_HALF_UP
import openpyxl

PARSER_VERSION = "1.0"

_MONTH_NAMES = frozenset([
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
])

_MONTH_ABBR_TO_FULL = {
    m[:3].lower(): m for m in _MONTH_NAMES
}


def _normalize_period_label(label: str) -> str:
    """
    Normalize a QB period column header to "Month YYYY" format.

    QB Balance Sheet exports sometimes use date-format headers:
      "December 2024"     → "December 2024"  (no change)
      "Dec 31, 2024"      → "December 2024"
      "December 31, 2024" → "December 2024"

    Returns the original label unchanged if it can't be normalized.
    """
    import re
    label = label.strip()

    # Already "Month YYYY"?
    parts = label.split()
    if len(parts) == 2 and parts[0] in _MONTH_NAMES and parts[1].isdigit() and len(parts[1]) == 4:
        return label

    # Strip leading "As of " prefix (e.g. "As of Dec 31, 2024")
    stripped = re.sub(r'^[Aa]s\s+of\s+', '', label).strip()

    # "Mon DD, YYYY" or "Month DD, YYYY" (with or without comma)
    m = re.match(r'^([A-Za-z]{3,9})\s+\d{1,2},?\s+(\d{4})$', stripped)
    if m:
        month_key = m.group(1).lower()[:3]
        year = m.group(2)
        full_month = _MONTH_ABBR_TO_FULL.get(month_key)
        if full_month:
            return f"{full_month} {year}"

    return label

# Rows that are calculated totals — skip them for mapping (would cause double-counting)
_PL_SKIP_NAMES = frozenset([
    "Gross Profit",
    "Net Operating Income",
    "Net Other Income",
    "Net Income",
])


def _to_cents(value) -> int:
    """Convert a float/int cell value to integer cents using decimal arithmetic."""
    if value is None or value == "":
        return 0
    try:
        d = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        return int(d * 100)
    except Exception:
        return 0


def _detect_report_type(cell_value) -> str:
    s = str(cell_value or "").strip()
    if "Profit and Loss" in s:
        return "profit_and_loss"
    if "Balance Sheet" in s:
        return "balance_sheet"
    return "unknown"


def _row_has_period_columns(row: tuple) -> bool:
    """Return True if any non-first cell in the row looks like a month/date period label."""
    import re
    for cell in row[1:]:
        if cell is None:
            continue
        label = str(cell).strip()
        # Matches "Month YYYY", "Mon DD, YYYY", "As of Mon DD, YYYY", date ranges, etc.
        if re.search(r'\b(January|February|March|April|May|June|July|August|September|October|November|December)\b', label):
            return True
    return False


def _find_header_row(all_rows: list) -> int:
    """
    Find the column header row — the row whose subsequent cells contain period labels.

    QB P&L exports label this row with "Distribution account" in col A.
    QB Balance Sheet exports may leave col A blank or use "Account".
    We detect by finding the first row that has at least one month-name period column.
    """
    # Prefer the exact "Distribution account" label (P&L)
    for i, row in enumerate(all_rows):
        if row and row[0] is not None and str(row[0]).strip() == "Distribution account":
            return i

    # Fall back: find the first row whose non-first cells look like period labels (Balance Sheet)
    for i, row in enumerate(all_rows):
        if row and _row_has_period_columns(row):
            return i

    raise ValueError("Could not find the column header row in this file. "
                     "Is this a QuickBooks export?")


def _parse_period_columns(header_row: tuple) -> list[tuple[str, int]]:
    """
    Return list of (period_label, column_index) for all month columns.
    Skips 'Total' and None columns.
    Period labels are normalized to "Month YYYY" where possible (e.g. "Dec 31, 2024" → "December 2024").
    """
    result = []
    for col_idx, cell in enumerate(header_row):
        if col_idx == 0:
            continue  # skip "Distribution account" label column
        if cell is None:
            continue
        label = str(cell).strip()
        if label.lower() == "total" or not label:
            continue
        result.append((_normalize_period_label(label), col_idx))
    return result


def _classify_row(account_name: str, raw_values: list, report_type: str = "") -> str:
    """
    Classify a row as: line_item | section_header | subtotal | skip
    raw_values: the original cell values (not converted to cents) for the period columns
    """
    name = account_name.strip()

    # Subtotals: "Total for X" prefix (named sub-group totals)
    if name.startswith("Total for "):
        return "subtotal"

    # QB section totals without "for": "Total Expenses", "Total Assets", "Total Payroll", etc.
    # These are always QB-generated calculated rows — never real accounts.
    if name.startswith("Total "):
        return "subtotal"

    # Known calculated P&L totals — only skip for P&L files.
    # "Net Income" in the Balance Sheet equity section IS a real line item we must capture.
    if report_type == "profit_and_loss" and name in _PL_SKIP_NAMES:
        return "subtotal"

    # Skip the footer line QB adds at the bottom
    if name.startswith("Accrual Basis") or name.startswith("Cash Basis"):
        return "skip"

    # Section header: all period cells are None (no values)
    if all(v is None for v in raw_values):
        return "section_header"

    return "line_item"


class _SectionState:
    """Tracks current section/subsection context as we walk through rows."""

    def __init__(self):
        self.section = ""
        self.subsection = ""

    def update_pl(self, name: str) -> None:
        """Update state based on a P&L section header name."""
        n = name.lower()
        if n == "income":
            self.section = "income"
            self.subsection = ""
        elif "cost of goods" in n:
            self.section = "cogs"
            self.subsection = ""
        elif n == "expenses":
            self.section = "expenses"
            self.subsection = "overhead"
        elif "payroll" in n:
            # Sub-section within expenses
            self.subsection = "payroll"
        elif "advertising" in n or "marketing" in n:
            self.subsection = "marketing"
        elif "depreciation" in n or "amortization" in n:
            self.subsection = "depreciation"
        elif "other income" in n:
            self.section = "other_income"
            self.subsection = ""
        elif "other expenses" in n:
            self.section = "other_expenses"
            self.subsection = ""
        # All other sub-group headers within expenses (e.g. "General business expenses",
        # "Vehicle Expenses", "Utilities") — do NOT change subsection.
        # They inherit the current subsection (overhead by default).

    def update_bs(self, name: str) -> None:
        """Update state based on a Balance Sheet section header name."""
        n = name.lower()
        if n == "assets":
            self.section = "assets"
            self.subsection = ""
        elif "current assets" in n and "other" not in n:
            self.subsection = "current_assets"
        elif "bank accounts" in n or "bank account" in n:
            self.subsection = "bank_accounts"
        elif n == "accounts receivable":
            self.subsection = "accounts_receivable"
        elif "other current assets" in n:
            self.subsection = "other_current_assets"
        elif "fixed assets" in n:
            self.subsection = "fixed_assets"
        elif "other assets" in n or ("long" in n and "assets" in n):
            self.subsection = "other_long_term_assets"
        elif "liabilities and equity" in n:
            self.section = "liabilities_equity"
            self.subsection = ""
        elif n == "liabilities":
            self.section = "liabilities"
            self.subsection = ""
        elif "current liabilities" in n and "other" not in n:
            self.subsection = "current_liabilities"
        elif "accounts payable" in n:
            self.subsection = "accounts_payable"
        elif "credit cards" in n or "credit card" in n:
            self.subsection = "credit_cards"
        elif "other current liabilities" in n:
            self.subsection = "other_current_liabilities"
        elif "long-term liabilities" in n or "long term liabilities" in n:
            self.subsection = "long_term_liabilities"
        elif n == "equity":
            self.section = "equity"
            self.subsection = "equity"
        elif "shareholders" in n or "stockholders" in n:
            self.subsection = "equity"
        # Sub-group headers within a known category (e.g. individual credit card names)
        # do NOT change subsection — they inherit from parent.

    def handle_subtotal_pl(self, name: str) -> None:
        """
        Reset subsection when exiting a named sub-section within Expenses.
        Called when we encounter a 'Total for X' subtotal row.
        """
        n = name.lower()
        if self.section == "expenses":
            if "payroll" in n or "advertising" in n or "marketing" in n or "depreciation" in n:
                self.subsection = "overhead"


def parse_file(file_bytes: bytes, filename: str) -> dict:
    """
    Parse a QuickBooks .xlsx export and return structured data.

    Returns:
        {
            "report_type": "profit_and_loss" | "balance_sheet",
            "company_name": str,
            "source_file": str,
            "periods_detected": ["January 2026", "February 2026", ...],
            "rows": [
                {
                    "account_name": str,
                    "row_type": "line_item",
                    "section": str,
                    "subsection": str,
                    "values": {"January 2026": int_cents, ...}
                },
                ...
            ],
            "parser_version": "1.0"
        }
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    if not all_rows:
        raise ValueError("File appears to be empty.")

    report_type = _detect_report_type(all_rows[0][0] if all_rows[0] else None)
    company_name = str(all_rows[1][0]).strip() if len(all_rows) > 1 and all_rows[1][0] else ""

    header_idx = _find_header_row(all_rows)
    period_cols = _parse_period_columns(all_rows[header_idx])

    if not period_cols:
        raise ValueError("No period columns found in this file.")

    periods = [p for p, _ in period_cols]
    col_map = {p: idx for p, idx in period_cols}

    state = _SectionState()
    rows = []

    for raw_row in all_rows[header_idx + 1:]:
        if not raw_row or raw_row[0] is None:
            continue

        account_name = str(raw_row[0]).strip()
        if not account_name:
            continue

        # Extract raw (unconverted) values for classification
        raw_values = [raw_row[idx] if idx < len(raw_row) else None for _, idx in period_cols]

        row_type = _classify_row(account_name, raw_values, report_type)

        if row_type == "skip":
            continue

        if row_type == "section_header":
            if report_type == "profit_and_loss":
                state.update_pl(account_name)
            else:
                state.update_bs(account_name)
            continue

        if row_type == "subtotal":
            if report_type == "profit_and_loss":
                state.handle_subtotal_pl(account_name)
            continue

        # line_item — convert values to cents and record
        values = {p: _to_cents(raw_row[idx] if idx < len(raw_row) else None)
                  for p, idx in period_cols}

        rows.append({
            "account_name": account_name,
            "row_type": "line_item",
            "section": state.section,
            "subsection": state.subsection,
            "values": values,
        })

    return {
        "report_type": report_type,
        "company_name": company_name,
        "source_file": filename,
        "periods_detected": periods,
        "rows": rows,
        "parser_version": PARSER_VERSION,
    }


def detect_report_type(file_bytes: bytes) -> str:
    """
    Read only the first cell to determine which QB report type this file is.
    Returns: "profit_and_loss" | "balance_sheet" | "invoices_by_month" | "unknown"
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(max_row=1, values_only=True))
    first_cell = str(rows[0][0] or "") if rows and rows[0] else ""
    if "Invoices by Month" in first_cell:
        return "invoices_by_month"
    if "Profit and Loss" in first_cell:
        return "profit_and_loss"
    if "Balance Sheet" in first_cell:
        return "balance_sheet"
    return "unknown"


def parse_invoice_report(file_bytes: bytes, filename: str) -> dict:
    """
    Parse a QuickBooks 'Invoices by Month' export.
    Returns invoice count per period, used as job_count.

    Format:
        Row 0: "Invoices by Month"
        Row 1: Company name
        Col A: Month label ("January 2026"), "Total for X", or None for detail rows
        Col B: Invoice date (not None for detail rows)
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    if not all_rows:
        raise ValueError("File appears to be empty.")

    company_name = str(all_rows[1][0]).strip() if len(all_rows) > 1 and all_rows[1][0] else ""

    from datetime import date as _date, datetime as _datetime

    current_month = None
    counts: dict[str, int] = {}

    for row in all_rows:
        if not row:
            continue
        col_a = row[0]
        col_b = row[1] if len(row) > 1 else None

        if col_a is not None:
            # Format A: date object in col_a → invoice detail row
            if isinstance(col_a, (_date, _datetime)) and current_month is not None:
                counts[current_month] += 1
                continue

            name = str(col_a).strip()
            # Skip totals, footer, and report-level header rows
            if name.startswith("Total") or name.startswith("TOTAL"):
                continue

            if col_b is None:
                # Month header: "December 2024"
                parts = name.split()
                if len(parts) == 2 and parts[0] in _MONTH_NAMES:
                    current_month = name
                    if current_month not in counts:
                        counts[current_month] = 0
            elif isinstance(col_b, (_date, _datetime)) and current_month is not None:
                # Format B: col_a = customer/invoice name, col_b = invoice date
                counts[current_month] += 1
        elif col_b is not None and current_month is not None:
            # Format C: col_a is None, col_b is an invoice date (datetime object or "MM/DD/YYYY" string)
            if isinstance(col_b, (_date, _datetime)) or (isinstance(col_b, str) and '/' in col_b):
                counts[current_month] += 1

    return {
        "report_type": "invoices_by_month",
        "company_name": company_name,
        "source_file": filename,
        "periods_detected": list(counts.keys()),
        "job_counts": counts,
    }
